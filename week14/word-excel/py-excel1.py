"""
Using the open-source package "OpenPyXL"
"""
from openpyxl import load_workbook

# Getting the excel file
workbook = load_workbook("test1.xlsx")

# test
# print(workbook)
# <openpyxl.workbook.workbook.Workbook object at 0x00000124B6F98BB0>

# .active property: Getting the currently active sheet or None
sheet = workbook.active
# Test:
print(sheet) # <Worksheet "Python">

# sheet.title="Python Class"
print(sheet.title) # Python

# Access the cell "A1":
print (sheet['A1']) # <Cell 'Python'.A1>

# Notice we can use small/capital => the same: a1 or A1
# Get the value of Cell "A1":
print (sheet['a1'].value) # Program Code

# Get the value of Cell "A4":
print (sheet['a4'].value) # Georgian@ILAC

# Add/change/overwrite a cell value:
sheet['A5'].value = "Dream Campus" # after saving our file:

# Or just mention the cell address when assign a value:
sheet['E1'] = "Time"
sheet['E2'] = "Wed 9AM"

"""
Notice that we do need to add .value when we want to access the cell value:
"""
print(sheet['E2'].value) # Wed 9AM

# change sheet name
sheet.title="Python Language"

"""
Notice that any change we make,
has to be saved in order to see the result in the excel file
"""

# workbook.save("test1.xlsx")
"""
Saving on the file when it's open will generate this error:
PermissionError: [Errno 13] Permission denied: 'test.xlsx'
"""

# printing the list (array) that contains all sheet names in the workbook:
print(workbook.sheetnames)
# ['Python Language', 'PHP', 'Java']

"""
To access other worksheets not only the active one: workbook['sheet_name']

- workbook['php']

Be advised that worksheet names are case-sensetive:
KeyError: 'Worksheet php does not exist.'
"""

# testing:
print (workbook['PHP']) # <Worksheet "PHP">


# creating a new sheet:
# create_sheet() method that accepts the sheet name:
workbook.create_sheet("C#")

# Remember that any change/update needs to be saved:
workbook.save('test1.xlsx')

# Testing:
print(workbook.sheetnames)
