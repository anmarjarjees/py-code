from openpyxl import  load_workbook

# get_column_letter function => will give us the column title out of the column index
from openpyxl.utils import get_column_letter

# Getting the excel file
wb = load_workbook("test2.xlsx")
sheet=wb.active

"""
With Excel:
- First row has the index value of 1
- First column also has the index value of 1
"""
# Looping through worksheet cells (access multiple cells):
# we have 5 rows: 1 to 6 in python range function

# (other programming languages) x = "A" + 1 => A1

# for (i=1; i<6; i++) :-)
for row in range(1, 6):
    # row=1, 2, 3, 4, 5
    # each row has two columns:
    """
    Notice here we need to get:
    1) A1, B1
    2) A2, B2
    3) A3, B3
    and so on...
    """
    for column in range(1,3): # 1 and 2 columns
        # column=1 and column=2 
        # print(sheet['A1'].value, sheet['B1'].value)
        
        # print(sheet[row,column].value)
        # Error: TypeError: expected string or bytes-like object

        """
        we can use a function from "openpyxl"
        called get_column_letter
        this function needs to be imported
        """
        letter = get_column_letter(column) # column=1 => A, column=2 => B
        
        # print(sheet[letter+row].value, sheet[letter+row].value)
        # Error: TypeError: can only concatenate str (not "int") to str 
        print(sheet[letter+str(row)].value)


